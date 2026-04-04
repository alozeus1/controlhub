"""Google Workspace publishing helpers for Agent artifacts."""

import csv
import io

from app.integrations.google_auth import build_drive_client, build_sheets_client


def publish_to_drive(file_bytes, filename, mime_type, folder_id, subject_email=None):
    """Upload a generated artifact to a Google Drive folder."""
    from googleapiclient.http import MediaInMemoryUpload

    service = build_drive_client(subject_email=subject_email)

    metadata = {
        "name": filename,
        "parents": [folder_id],
    }
    media = MediaInMemoryUpload(file_bytes, mimetype=mime_type, resumable=False)

    response = (
        service.files()
        .create(body=metadata, media_body=media, fields="id,name,webViewLink")
        .execute()
    )
    return {
        "drive_file_id": response.get("id"),
        "name": response.get("name"),
        "web_view_link": response.get("webViewLink"),
    }


def _parse_rows_from_artifact(file_bytes, mime_type):
    if mime_type in {"text/csv", "application/csv"}:
        text = file_bytes.decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        return [list(row) for row in reader]

    if mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        from openpyxl import load_workbook

        workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
        sheet = workbook.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([value for value in row])
        return rows

    # Fallback: one-column data
    text = file_bytes.decode("utf-8")
    return [[line] for line in text.splitlines()]


def publish_to_sheet(file_bytes, mime_type, spreadsheet_id, sheet_name, a1_range, mode="overwrite", subject_email=None):
    """Publish artifact tabular data into Google Sheets."""
    if mode not in {"overwrite", "append"}:
        raise ValueError("mode must be overwrite or append")

    service = build_sheets_client(subject_email=subject_email)
    rows = _parse_rows_from_artifact(file_bytes, mime_type)

    target_range = f"{sheet_name}!{a1_range}" if sheet_name else a1_range

    if mode == "overwrite":
        result = (
            service.spreadsheets()
            .values()
            .update(
                spreadsheetId=spreadsheet_id,
                range=target_range,
                valueInputOption="RAW",
                body={"values": rows},
            )
            .execute()
        )
        return {
            "updated_range": result.get("updatedRange"),
            "updated_rows": result.get("updatedRows"),
            "updated_columns": result.get("updatedColumns"),
            "updated_cells": result.get("updatedCells"),
        }

    result = (
        service.spreadsheets()
        .values()
        .append(
            spreadsheetId=spreadsheet_id,
            range=target_range,
            valueInputOption="RAW",
            insertDataOption="INSERT_ROWS",
            body={"values": rows},
        )
        .execute()
    )
    updates = result.get("updates", {})
    return {
        "updated_range": updates.get("updatedRange"),
        "updated_rows": updates.get("updatedRows"),
        "updated_columns": updates.get("updatedColumns"),
        "updated_cells": updates.get("updatedCells"),
    }
